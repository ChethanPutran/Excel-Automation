import openpyxl as xl
from openpyxl.chart import BarChart,Reference

#Loading the excel workbook
wb = xl.load_workbook('Book1.xlsx')

#Accessing thr first sheet
sheet = wb['Sheet1']
cell = sheet['a1']
#cell = sheet1.cell(1,1)
# print(cell.value)
# print(sheet.max_row)

#Manipulating the third column values
for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    # print(cell.value)
    corrected_value = cell.value * 2
    corrected_value_cell = sheet.cell(row,4)
    corrected_value_cell.value = corrected_value

   
values = Reference(sheet,min_row = 2 ,max_row = sheet.max_row, min_col = 4,max_col = 4) 

#Adding BarChart to the excel sheet
chart = BarChart()
chart.add_data(values)

#Displaying the BarChart at 5th column
sheet.add_chart(chart,'e1')

wb.save('Book2.xlsx')